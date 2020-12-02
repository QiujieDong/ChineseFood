#coding=utf-8
import re
import requests
from urllib import error
from bs4 import BeautifulSoup
import os
import openpyxl
from PIL import Image
import imghdr

num = 0
numPicture = 0
file = ''
List = []


def Find(url):
    global List
    print('正在检测图片总数，请稍等.....')
    t = 0
    i = 1
    s = 0
    while t < 1000:
        Url = url + str(t) + '&gsm=8c'
        try:
            Result = requests.get(Url, timeout=7)
        except BaseException:
            t = t+60
            continue
        else:
            result = Result.text
            pic_url = re.findall('"objURL":"(.*?)",', result, re.S)  # 先利用正则表达式找到图片url
            s += len(pic_url)
            if len(pic_url) == 0:
                break
            else:
                List.append(pic_url)
                t = t + 60
    return s


def recommend(url):
    Re = []
    try:
        html = requests.get(url)
    except error.HTTPError as e:
        return
    else:
        html.encoding = 'utf-8'
        bsObj = BeautifulSoup(html.text, 'html.parser')
        div = bsObj.find('div', id='topRS')
        if div is not None:
            listA = div.findAll('a')
            for i in listA:
                if i is not None:
                    Re.append(i.get_text())
        return Re


def dowmloadPicture(html, keyword):
    global num
    pic_url = re.findall('"objURL":"(.*?)",', html, re.S)  # 先利用正则表达式找到图片url
    print('找到关键词:' + keyword + '的图片，即将开始下载图片...')
    for each in pic_url:
        print('正在下载第' + str(num + 1) + '张图片，图片地址:' + str(each))
        try:
            if each is not None:
                pic = requests.get(each, timeout=10)
            else:
                continue
        except BaseException:
            print('错误，当前图片无法下载')
            continue
        else:
            string = file + keyword + '_' + str(num) + '.jpg'
            fp = open(string, 'wb')
            fp.write(pic.content)
            fp.close()
            num += 1
        if num >= numPicture:
            return


if __name__ == '__main__':  # 主函数入口
    # 打开excel文件,获取工作簿对象
    wb = openpyxl.load_workbook('./food.xlsx')
    # 从表单中获取单元格的内容
    ws = wb.active  # 当前活跃的表单
    

    for i in range(int(ws.max_row/2)+9, int(ws.max_row),1): 
        word = (ws.cell(row=i, column=2).value)
        url = 'http://image.baidu.com/search/flip?tn=baiduimage&ie=utf-8&word='+word+'&ct=201326592&v=flip'
        numPicture = 1000
        file = './' + word + '/'
        y = os.path.exists(file)
        if y == 1:
            print('该文件已存在，请重新输入')
            file = input('请建立一个存储图片的文件夹，输入文件夹名称即可')
            os.mkdir(file)
        else:
            os.mkdir(file)
        t = 0
        while t < numPicture:
            try:
                url = url + '&pn='+str(t) + '&gsm=8c'
                result = requests.get(url, timeout=1200)
            except error.HTTPError as e:
                print('网络错误，请调整网络后重试')
            else:
                dowmloadPicture(result.text, word)
            finally:
                t = t+20
        t = 0
        num =  0
        print('当前搜索结束，感谢使用')
